#bulk chrome extension IDs validation python script
from apiclient.discovery import build
import openpyxl

api_key = "xxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
res = []
resource = build("customsearch", 'v1', developerKey=api_key).cse()
path = "C:\\Users\\nballa\\Desktop\\python\\test\\googlesearch\\url.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
item = []
for i in range(1, m_row + 1):
	k = sheet_obj.cell(row = i, column = 1)
	#m = "('malware' OR 'virus' OR 'adware' OR 'exploit' OR 'vulnerabilities')"
	z = k.value 
	x= z #+ m
	
	#print(z)
	result = resource.list(q=x, cx='03b7e7f18d0184522').execute()
	for item in result['items']:
		#print(z, item['title'], item['link'])
		res.append([z, item['title'], item['link']])
		break
wb = openpyxl.Workbook()
sheet = wb.active
sheet_title = "chromeextensionvalidation"
for i in res:
	sheet.append(i)
wb.save("C:\\Users\\nballa\\Desktop\\python\\test\\googlesearch\\chromeextensionvalidation.xlsx")

#in progress#
