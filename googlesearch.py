from apiclient.discovery import build
import openpyxl

api_key = "XXXXXXXXXXXXXXXX"

resource = build("customsearch", 'v1', developerKey=api_key).cse()
path = "C:\\Users\\url.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
m = "malware"
for i in range(1, m_row + 1):
	k = sheet_obj.cell(row = i, column = 1)
	m = "('malware' OR 'virus' OR 'adware' OR 'exploit' OR 'vulnerabilities')"
	z = k.value 
	x= z + m
	
	print(z)
	result = resource.list(q=x, cx='03b7e7f18d0184522').execute()
	for item in result['items']:
		print(item['title'], item['link'])
		
# in progress#   
